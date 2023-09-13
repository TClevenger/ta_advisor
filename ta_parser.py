# Quick and dirty AWS TrustedAdvisor parser - removes sheets where the status
# is "ok" or "not_available" and leaves sheets with warnings and errors.  This
# does not create backups or prompt for confirmation so use as your own risk.
# Tim Clevenger 2023 (https://github.com/TClevenger)
#
# openpyxl required:  pip3 install openpyxl
#

from openpyxl import load_workbook
import sys
import warnings

# Suppress 'using default style' warning
warnings.simplefilter("ignore")

if len(sys.argv) == 1:
  print("usage: python3 ta_parser.py FILE")
  sys.exit(1)

delete_sheets = []
wbname = sys.argv[1]
workbook = load_workbook(wbname)
changes = False

for x in workbook.sheetnames:
  sheet = workbook[x]
  if str(sheet["A4"].value) == "Status: ok" or str(sheet["A4"].value) == "Status: not_available":
    delete_sheets.append(sheet)
  else:
    print("Sheet name: " + str(sheet.title) + " will be kept.")
    changes = True

if not changes:
  print("No warnings or errors found in report; no changes made to original file.")
  sys.exit(0)

for x in delete_sheets:
  print("Deleting sheet " + str(x))
  workbook.remove_sheet(x)

workbook.save(filename = wbname)
print("File has been updated.")

sys.exit(0)
